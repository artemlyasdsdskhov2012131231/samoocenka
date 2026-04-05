import csv
import ctypes
import json
import os
import random
import subprocess
import sys
import traceback
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path

import flet as ft
from storage import get_results_dir, get_logs_dir, get_exports_dir, get_backups_dir, get_platform_name

# Обработка необязательных зависимостей
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as e:
    print(f"⚠️  WARNING: openpyxl not available: {e}")
    print("   Install with: pip install openpyxl")
    Workbook = None
    Alignment = Font = PatternFill = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
except ImportError as e:
    print(f"⚠️  WARNING: reportlab not available: {e}")
    print("   Install with: pip install reportlab Pillow")
    colors = A4 = ParagraphStyle = getSampleStyleSheet = mm = None
    pdfmetrics = TTFont = None
    Paragraph = SimpleDocTemplate = Spacer = Table = TableStyle = None


APP_TITLE = "Самооценка состояния"
APP_SUBTITLE = "Экспресс-анкета психологического самочувствия"

SENSATIONS = [
    "Усталость",
    "Невнимательность",
    "Хорошее настроение",
    "Плохое самочувствие",
    "Уверенность в себе",
    "Безразличие",
    "Плохое настроение",
    "Хорошее общее самочувствие",
    "Спокойствие",
    "Неуверенность",
]

COMPLAINTS = [
    "Головокружение",
    "Головная боль",
    "Чувство тяжести в голове",
    "Сонливость",
    "Ощущение жара",
    "Жажда",
    "Ощущение голода",
    "Необычный вкус во рту",
    "Затруднение дыхания",
    "Сердцебиение",
    "Сухость во рту",
    "Боли в груди",
    "Слюнотечение",
    "Потливость",
    "Подташнивание",
    "Мышечная слабость",
    "Как будто принял лекарство",
    "Неприятные ощущения в животе",
]

POSITIVE_SENSATIONS = {
    "Хорошее настроение",
    "Уверенность в себе",
    "Хорошее общее самочувствие",
    "Спокойствие",
}
NEGATIVE_SENSATIONS = [item for item in SENSATIONS if item not in POSITIVE_SENSATIONS]
SCRIPT_PATH = Path(globals().get("__file__", Path.cwd())).resolve()
BASE_DIR = (
    Path(sys.executable).resolve().parent
    if getattr(sys, "frozen", False)
    else (SCRIPT_PATH.parent if SCRIPT_PATH.is_file() else SCRIPT_PATH)
)

# Кроссплатформные пути к директориям
RESULTS_DIR = get_results_dir()
LOGS_DIR = get_logs_dir()
EXPORTS_DIR = get_exports_dir()
BACKUPS_DIR = get_backups_dir()


# Load questions bank from JSON
def load_test_questions():
    """Load questions from test_questions.json for randomization"""
    try:
        questions_file = BASE_DIR / "test_questions.json"
        if questions_file.exists():
            with open(questions_file, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"DEBUG: Could not load test_questions.json: {e}")
    return None


QUESTIONS_BANK = load_test_questions()


def get_random_questions(category: str, count: int = None) -> list:
    """Get random questions from a category in the questions bank"""
    if not QUESTIONS_BANK or category not in QUESTIONS_BANK:
        return []
    
    category_data = QUESTIONS_BANK.get(category, {})
    questions = category_data.get("questions", [])
    
    if not questions:
        return []
    
    # Determine how many questions to select
    min_q = category_data.get("min_questions", 1)
    max_q = category_data.get("max_questions", len(questions))
    
    if count is None:
        count = min(random.randint(min_q, max_q), len(questions))
    else:
        count = min(count, len(questions))
    
    # Select random questions without replacement
    selected = random.sample(questions, count)
    return selected


def log_platform_info():
    """Логирует информацию о платформе."""
    try:
        platform = get_platform_name()
        print(f"DEBUG: Platform detected: {platform}")
    except Exception as e:
        print(f"DEBUG: Platform detection error: {e}")


def sanitize_filename_part(value: str, fallback: str) -> str:
    cleaned = "".join(ch for ch in (value or "").strip() if ch not in '<>:"/\\|?*')
    cleaned = "_".join(cleaned.split())
    return cleaned or fallback


def extract_surname(fio: str) -> str:
    parts = [part for part in (fio or "").replace(",", " ").split() if part]
    surname = parts[0] if parts else "Без_фамилии"
    return sanitize_filename_part(surname, "Без_фамилии")


RESULT_TABLE_HEADERS = [
    "Дата",
    "Время",
    "Фамилия",
    "Ф.И.О.",
    "Звание",
    "Подразделение",
    "Вид дежурства",
    "Состояние",
    "Индекс состояния",
    "Позитивные маркеры",
    "Негативные маркеры",
    "Количество жалоб",
    "Краткий анализ",
    "Резюме",
    "Рекомендация",
    "Жалобы",
]


def build_daily_results_path(user_data) -> Path:
    date_part = sanitize_filename_part(
        user_data.get("date", datetime.now().strftime("%d.%m.%Y")),
        datetime.now().strftime("%d.%m.%Y"),
    )
    return RESULTS_DIR / f"{date_part}.csv"


def build_result_row(user_data, result, complaints_list):
    return {
        "Дата": user_data.get("date", ""),
        "Время": user_data.get("time", ""),
        "Фамилия": extract_surname(user_data.get("fio", "")),
        "Ф.И.О.": user_data.get("fio", ""),
        "Звание": user_data.get("rank", ""),
        "Подразделение": user_data.get("unit", ""),
        "Вид дежурства": user_data.get("duty", ""),
        "Состояние": result["status_text"],
        "Индекс состояния": result["final_score"],
        "Позитивные маркеры": result["positive_score"],
        "Негативные маркеры": result["negative_score"],
        "Количество жалоб": result["complaints_count"],
        "Краткий анализ": " | ".join(result["details"]),
        "Резюме": result["summary"],
        "Рекомендация": result["recommendation"],
        "Жалобы": ", ".join(complaints_list) if complaints_list else "Жалобы не отмечены",
    }


def ensure_results_file_schema(file_path: Path):
    if not file_path.exists() or file_path.stat().st_size == 0:
        return

    with file_path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file, delimiter=";")
        existing_headers = reader.fieldnames or []
        if existing_headers == RESULT_TABLE_HEADERS:
            return
        existing_rows = list(reader)

    normalized_rows = []
    for row in existing_rows:
        normalized_rows.append({header: row.get(header, "") for header in RESULT_TABLE_HEADERS})

    with file_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=RESULT_TABLE_HEADERS,
            delimiter=";",
        )
        writer.writeheader()
        writer.writerows(normalized_rows)



def save_result_to_table(user_data, result, complaints_list):
    try:
        RESULTS_DIR.mkdir(parents=True, exist_ok=True)
        file_path = build_daily_results_path(user_data)
        write_header = not file_path.exists() or file_path.stat().st_size == 0

        if not write_header:
            ensure_results_file_schema(file_path)

        with file_path.open("a", encoding="utf-8-sig", newline="") as file:
            writer = csv.DictWriter(
                file,
                fieldnames=RESULT_TABLE_HEADERS,
                delimiter=";",
            )
            if write_header:
                writer.writeheader()
            writer.writerow(build_result_row(user_data, result, complaints_list))

        try:
            create_results_backup("auto")
        except OSError:
            pass

        return file_path, None
    except OSError as exc:
        return None, str(exc)


def ensure_storage_dirs():
    """Создать директории хранилища с обработкой ошибок"""
    try:
        for directory in (RESULTS_DIR, LOGS_DIR, EXPORTS_DIR, BACKUPS_DIR):
            directory.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        print(f"ERROR: Failed to create storage dirs: {e}")
        traceback.print_exc()


def parse_score_value(value) -> int:
    try:
        return int(str(value).split("/")[0].strip())
    except (TypeError, ValueError):
        return 0


def get_current_date_text() -> str:
    return datetime.now().strftime("%d.%m.%Y")


def get_current_time_text() -> str:
    return datetime.now().strftime("%H:%M:%S")


def parse_date_value(value: str) -> datetime:
    for fmt in ("%d.%m.%Y", "%d.%m.%Y %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime((value or "").strip(), fmt)
        except ValueError:
            continue
    return datetime.min


def normalize_status_name(status_text: str) -> str:
    if not status_text:
        return "Не определено"
    return status_text.split("—")[0].strip()


def get_status_color(status_text: str) -> str:
    normalized = normalize_status_name(status_text)
    if normalized.startswith("Зелёный"):
        return ft.Colors.GREEN_600
    if normalized.startswith("Жёлтый"):
        return ft.Colors.AMBER_700
    if normalized.startswith("Оранжевый"):
        return ft.Colors.ORANGE_700
    if normalized.startswith("Красный"):
        return ft.Colors.RED_600
    return ft.Colors.BLUE_GREY_600


def get_status_icon(status_text: str) -> str:
    normalized = normalize_status_name(status_text)
    if normalized.startswith("Зелёный"):
        return ft.Icons.CHECK_CIRCLE_ROUNDED
    if normalized.startswith("Жёлтый"):
        return ft.Icons.INFO_ROUNDED
    if normalized.startswith("Оранжевый"):
        return ft.Icons.WARNING_AMBER_ROUNDED
    if normalized.startswith("Красный"):
        return ft.Icons.ERROR_ROUNDED
    return ft.Icons.INSIGHTS_ROUNDED


def load_history_rows():
    """Load history with robust error handling for Android storage issues."""
    rows = []
    try:
        if not RESULTS_DIR.exists():
            return rows
    except (NameError, AttributeError) as e:
        print(f"ERROR: RESULTS_DIR not accessible: {e}")
        return rows

    try:
        for file_path in sorted(RESULTS_DIR.glob("*.csv"), reverse=True):
            try:
                with file_path.open("r", encoding="utf-8-sig", newline="") as file:
                    reader = csv.DictReader(file, delimiter=";")
                    for row in reader:
                        if any((value or "").strip() for value in row.values()):
                            rows.append(row)
            except OSError as e:
                print(f"DEBUG: Could not read {file_path}: {e}")
                continue
            except Exception as e:
                print(f"WARNING: Error processing file {file_path}: {e}")
                continue
    except (NameError, Exception) as e:
        print(f"ERROR: Error loading history: {e}")
        return rows

    try:
        rows.sort(key=lambda row: parse_date_value(row.get("Дата", "")), reverse=True)
    except Exception as e:
        print(f"WARNING: Could not sort history: {e}")
    
    return rows


def build_export_basename(user_data) -> str:
    surname = extract_surname(user_data.get("fio", ""))
    date_part = sanitize_filename_part(
        user_data.get("date", datetime.now().strftime("%d.%m.%Y")),
        datetime.now().strftime("%d.%m.%Y"),
    )
    time_part = sanitize_filename_part(user_data.get("time", get_current_time_text()), get_current_time_text())
    return f"{date_part}_{time_part}_{surname}"


def cleanup_old_backups(max_files: int = 25):
    backup_files = sorted(BACKUPS_DIR.glob("*.zip"), key=lambda path: path.stat().st_mtime, reverse=True)
    for stale_file in backup_files[max_files:]:
        try:
            stale_file.unlink()
        except OSError:
            continue


def create_results_backup(tag: str = "manual"):
    ensure_storage_dirs()
    archive_name = f"backup_{tag}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    archive_path = BACKUPS_DIR / archive_name

    with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        root_dir = RESULTS_DIR.parent
        for folder in (RESULTS_DIR, EXPORTS_DIR, LOGS_DIR):
            if not folder.exists():
                continue
            for file_path in folder.rglob("*"):
                if file_path.is_file():
                    archive.write(file_path, arcname=file_path.relative_to(root_dir))

    cleanup_old_backups()
    return archive_path, None


def export_history_to_excel(history_rows):
    if Workbook is None:
        return None, "❌ Пакет openpyxl недоступен.\nУстановите: pip install openpyxl"
    if not history_rows:
        return None, "История результатов пока пуста."

    try:
        ensure_storage_dirs()
    except (NameError, Exception) as e:
        return None, f"❌ Ошибка инициализации хранилища:\n{str(e)}"

    try:
        file_path = EXPORTS_DIR / f"history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Сводка"

        title_font = Font(size=16, bold=True, color="B42318")
        header_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="B42318")

        summary_sheet["A1"] = APP_TITLE
        summary_sheet["A2"] = "Сводка по истории результатов"
        summary_sheet["A1"].font = title_font
        summary_sheet["A2"].font = Font(size=11, italic=True, color="666666")

        counts = Counter(normalize_status_name(row.get("Состояние", "")) for row in history_rows)
        summary_rows = [
            ("Всего записей", len(history_rows)),
            ("Зелёный", counts.get("Зелёный", 0)),
            ("Жёлтый", counts.get("Жёлтый", 0)),
            ("Оранжевый", counts.get("Оранжевый", 0)),
            ("Красный", counts.get("Красный", 0)),
        ]
        for row_index, (label, value) in enumerate(summary_rows, start=4):
            summary_sheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
            summary_sheet.cell(row=row_index, column=2, value=value)

        daily_scores = {}
        daily_counts = {}
        for row in history_rows:
            date_value = row.get("Дата", "")
            if not date_value:
                continue
            daily_scores[date_value] = daily_scores.get(date_value, 0) + parse_score_value(row.get("Индекс состояния", 0))
            daily_counts[date_value] = daily_counts.get(date_value, 0) + 1

        summary_sheet["D3"] = "Средний индекс по датам"
        summary_sheet["D3"].font = Font(bold=True)
        summary_sheet["D4"] = "Дата"
        summary_sheet["E4"] = "Средний индекс"
        for cell in (summary_sheet["D4"], summary_sheet["E4"]):
            cell.font = header_font
            cell.fill = fill
        for row_index, date_value in enumerate(sorted(daily_scores.keys(), key=parse_date_value), start=5):
            average_score = round(daily_scores[date_value] / max(daily_counts[date_value], 1), 2)
            summary_sheet.cell(row=row_index, column=4, value=date_value)
            summary_sheet.cell(row=row_index, column=5, value=average_score)

        summary_sheet.column_dimensions["A"].width = 28
        summary_sheet.column_dimensions["B"].width = 18
        summary_sheet.column_dimensions["D"].width = 18
        summary_sheet.column_dimensions["E"].width = 18

        history_sheet = workbook.create_sheet("История")
        for column, header in enumerate(RESULT_TABLE_HEADERS, start=1):
            cell = history_sheet.cell(row=1, column=column, value=header)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in history_rows:
            history_sheet.append([row.get(header, "") for header in RESULT_TABLE_HEADERS])
        for column_cells in history_sheet.columns:
            history_sheet.column_dimensions[column_cells[0].column_letter].width = 20

        workbook.save(file_path)
        return file_path, None
    except OSError as exc:
        return None, str(exc)


def build_official_conclusion(user_data, result, complaints_list) -> str:
    complaints_text = ", ".join(complaints_list) if complaints_list else "жалобы не отмечены"
    return (
        f"{user_data.get('date', '')} в {user_data.get('time', '')} обследуемый {user_data.get('fio', '')} "
        f"({user_data.get('rank', '')}, {user_data.get('unit', '')}) прошёл экспресс-оценку состояния. "
        f"Итог: {result.get('status_text', '').lower()}. "
        f"Индекс состояния — {result.get('final_score', 0)} из 20. "
        f"Жалобы: {complaints_text}. "
        f"Заключение: {result.get('summary', '')} "
        f"Рекомендация: {result.get('recommendation', '')}"
    )


def export_result_to_excel(user_data, result, complaints_list):
    if Workbook is None:
        return None, "Пакет openpyxl недоступен."

    try:
        ensure_storage_dirs()
        file_path = EXPORTS_DIR / f"{build_export_basename(user_data)}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Результат"

        title_font = Font(size=16, bold=True, color="B42318")
        header_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="B42318")

        sheet["A1"] = APP_TITLE
        sheet["A2"] = APP_SUBTITLE
        sheet["A1"].font = title_font
        sheet["A2"].font = Font(size=10, italic=True, color="666666")

        info_rows = [
            ("Дата", user_data.get("date", "")),
            ("Время прохождения", user_data.get("time", "")),
            ("Ф.И.О.", user_data.get("fio", "")),
            ("Звание", user_data.get("rank", "")),
            ("Подразделение", user_data.get("unit", "")),
            ("Вид дежурства", user_data.get("duty", "")),
            ("Состояние", result["status_text"]),
            ("Индекс состояния", f"{result['final_score']} / 20"),
            ("Позитивные маркеры", result["positive_score"]),
            ("Негативные маркеры", result["negative_score"]),
            ("Количество жалоб", result["complaints_count"]),
            ("Резюме", result["summary"]),
            ("Рекомендация", result["recommendation"]),
            ("Жалобы", ", ".join(complaints_list) if complaints_list else "Жалобы не отмечены"),
        ]

        for row_index, (label, value) in enumerate(info_rows, start=4):
            sheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
            sheet.cell(row=row_index, column=2, value=value)
            sheet.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True, vertical="top")

        sheet.column_dimensions["A"].width = 26
        sheet.column_dimensions["B"].width = 72

        data_sheet = workbook.create_sheet("Данные")
        for column, header in enumerate(RESULT_TABLE_HEADERS, start=1):
            cell = data_sheet.cell(row=1, column=column, value=header)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        result_row = build_result_row(user_data, result, complaints_list)
        data_sheet.append([result_row.get(header, "") for header in RESULT_TABLE_HEADERS])
        for column_cells in data_sheet.columns:
            data_sheet.column_dimensions[column_cells[0].column_letter].width = 22

        workbook.save(file_path)
        return file_path, None
    except OSError as exc:
        return None, str(exc)


def get_pdf_font_name() -> str:
    if pdfmetrics is None or TTFont is None:
        return "Helvetica"

    for font_name, font_path in (
        ("AppArial", Path("C:/Windows/Fonts/arial.ttf")),
        ("AppSegoeUI", Path("C:/Windows/Fonts/segoeui.ttf")),
    ):
        if font_path.exists():
            try:
                pdfmetrics.getFont(font_name)
            except KeyError:
                pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
            return font_name

    return "Helvetica"


def export_result_to_pdf(user_data, result, complaints_list):
    if SimpleDocTemplate is None or colors is None or mm is None:
        return None, "❌ Пакет reportlab недоступен.\nУстановите: pip install reportlab Pillow"

    try:
        ensure_storage_dirs()
    except (NameError, Exception) as e:
        return None, f"❌ Ошибка инициализации хранилища:\n{str(e)}"

    try:
        file_path = EXPORTS_DIR / f"{build_export_basename(user_data)}.pdf"
        font_name = get_pdf_font_name()
        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=A4,
            leftMargin=16 * mm,
            rightMargin=16 * mm,
            topMargin=14 * mm,
            bottomMargin=14 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "TitleStyle",
            parent=styles["Title"],
            fontName=font_name,
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#B42318"),
        )
        heading_style = ParagraphStyle(
            "HeadingStyle",
            parent=styles["Heading2"],
            fontName=font_name,
            fontSize=13,
            leading=16,
            textColor=colors.HexColor("#101828"),
            spaceAfter=6,
        )
        body_style = ParagraphStyle(
            "BodyStyle",
            parent=styles["BodyText"],
            fontName=font_name,
            fontSize=10,
            leading=14,
        )

        complaints_text = ", ".join(complaints_list) if complaints_list else "Жалобы не отмечены"
        details_text = "<br/>".join(f"• {item}" for item in result["details"])
        info_table = Table(
            [
                ["Дата", user_data.get("date", "")],
                ["Время прохождения", user_data.get("time", "")],
                ["Ф.И.О.", user_data.get("fio", "")],
                ["Звание", user_data.get("rank", "")],
                ["Подразделение", user_data.get("unit", "")],
                ["Вид дежурства", user_data.get("duty", "")],
                ["Состояние", result["status_text"]],
                ["Индекс состояния", f"{result['final_score']} / 20"],
                ["Количество жалоб", str(result["complaints_count"])],
            ],
            colWidths=[48 * mm, 120 * mm],
        )
        info_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#FEE4E2")),
                    ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#7A271A")),
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#FECACA")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("PADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )

        story = [
            Paragraph(APP_TITLE, title_style),
            Paragraph(APP_SUBTITLE, body_style),
            Spacer(1, 5 * mm),
            info_table,
            Spacer(1, 5 * mm),
            Paragraph("Резюме", heading_style),
            Paragraph(result["summary"], body_style),
            Spacer(1, 3 * mm),
            Paragraph("Краткий анализ", heading_style),
            Paragraph(details_text, body_style),
            Spacer(1, 3 * mm),
            Paragraph("Рекомендация", heading_style),
            Paragraph(result["recommendation"], body_style),
            Spacer(1, 3 * mm),
            Paragraph("Отмеченные жалобы", heading_style),
            Paragraph(complaints_text, body_style),
        ]

        doc.build(story)
        return file_path, None
    except OSError as exc:
        return None, str(exc)


def generate_daily_report(user_data):
    """Generate daily duty report with 4+ test sessions for the day."""
    if not user_data or not user_data.get("fio"):
        return None, "Требуются данные пользователя."
    
    # Extract surname from FIO
    fio = user_data.get("fio", "").strip()
    surname = extract_surname(fio)
    current_date = user_data.get("date", get_current_date_text())
    
    # Load all results for current date
    all_results = load_history_rows()
    
    # Filter results for same person and date
    daily_results = []
    for row in all_results:
        if (row.get("Дата") == current_date and 
            extract_surname(row.get("Ф.И.О.", "")) == surname):
            daily_results.append(row)
    
    # Check if we have at least 4 results
    if len(daily_results) < 4:
        return None, f"Требуется минимум 4 проведения теста. Текущий результат: {len(daily_results)}/4"
    
    # Calculate statistics
    scores = []
    all_complaints = []
    statuses = []
    
    for result in daily_results:
        try:
            score = float(result.get("Индекс состояния", "0"))
            scores.append(score)
        except (ValueError, TypeError):
            pass
        
        complaints_str = result.get("Количество жалоб", "0")
        try:
            complaint_count = int(complaints_str) if complaints_str else 0
        except (ValueError, TypeError):
            complaint_count = 0
        
        # Extract complaints (they are in a column - we'll just note that complaints were present)
        status = result.get("Состояние", "")
        statuses.append(normalize_status_name(status))
    
    avg_score = sum(scores) / len(scores) if scores else 0
    
    # Determine daily status
    avg_status = ""
    if avg_score >= 15:
        avg_status = "Зелёный"
        daily_assessment = "День прошёл успешно. Уровень состояния в норме."
    elif avg_score >= 11:
        avg_status = "Жёлтый"
        daily_assessment = "День прошёл с некоторыми трудностями. Рекомендуется повышенное внимание к состоянию."
    elif avg_score >= 6:
        avg_status = "Оранжевый"
        daily_assessment = "День был напряжённым. Рекомендуется отдых и восстановление."
    else:
        avg_status = "Красный"
        daily_assessment = "Дежурство было очень сложным. Срочно требуется помощь психолога."
    
    return {
        "date": current_date,
        "fio": fio,
        "surname": surname,
        "test_count": len(daily_results),
        "avg_score": avg_score,
        "status": avg_status,
        "assessment": daily_assessment,
        "daily_results": daily_results,
        "statuses": statuses,
    }, None


def export_daily_report(daily_report_data):
    """Export daily duty report to plain text file (.txt with UTF-8).
    
    Exports only readable text - NO binary or Excel data.
    Opens in notepad without any corruption.
    """
    if not daily_report_data or not isinstance(daily_report_data, dict):
        return None, "Некорректные данные отчёта."
    
    try:
        ensure_storage_dirs()
        
        # Create filename with date and surname - ONLY TEXT
        date_str = daily_report_data.get("date", "").replace(".", "")
        surname = sanitize_filename_part(daily_report_data.get("surname", ""), "Report")
        filename = f"СУТОЧНЫЙ_ОТЧЕТ_{date_str}_{surname}.txt"
        file_path = EXPORTS_DIR / filename
        
        # Build report content as PURE TEXT ONLY
        content = []
        
        # Header
        content.append("=" * 70)
        content.append("СУТОЧНЫЙ ОТЧЕТ О СОСТОЯНИИ")
        content.append("=" * 70)
        content.append("")
        
        # Basic info - SANITIZE ALL TEXT
        fio_clean = str(daily_report_data.get('fio', '')).strip()
        # Remove any non-printable characters
        fio_clean = ''.join(c for c in fio_clean if ord(c) >= 32 or c in '\n\t\r')
        
        date_clean = str(daily_report_data.get('date', '')).strip()
        date_clean = ''.join(c for c in date_clean if ord(c) >= 32 or c in '\n\t\r')
        
        content.append(f"Дата: {date_clean}")
        content.append(f"Ф.И.О.: {fio_clean}")
        content.append("")
        
        # Summary section
        content.append("-" * 70)
        content.append("СВОДКА ПО ДЕЖУРСТВУ")
        content.append("-" * 70)
        
        test_count = daily_report_data.get('test_count', 0)
        content.append(f"Всего проведений теста: {test_count}")
        
        avg_score = daily_report_data.get('avg_score', 0)
        content.append(f"Средний индекс состояния: {avg_score:.1f} / 20")
        
        status_clean = str(daily_report_data.get('status', '')).strip()
        status_clean = ''.join(c for c in status_clean if ord(c) >= 32 or c in '\n\t\r')
        content.append(f"Общее состояние: {status_clean}")
        content.append("")
        
        content.append("ЗАКЛЮЧЕНИЕ:")
        assessment_clean = str(daily_report_data.get('assessment', '')).strip()
        assessment_clean = ''.join(c for c in assessment_clean if ord(c) >= 32 or c in '\n\t\r')
        content.append(assessment_clean)
        content.append("")
        
        # Results table
        content.append("-" * 70)
        content.append("РЕЗУЛЬТАТЫ ВСЕХ ПРОВЕДЕНИЙ")
        content.append("-" * 70)
        content.append("")
        
        # Table header
        content.append("№ | Время      | Состояние                        | Индекс | Жалобы | Подразд.")
        content.append("-" * 70)
        
        # Add results to table - SANITIZE ALL DATA
        for idx, result in enumerate(daily_report_data.get('daily_results', []), 1):
            # Extract and sanitize each field
            time_str = str(result.get("Время", "")).strip()
            time_str = ''.join(c for c in time_str if ord(c) >= 32 or c in '\n\t\r')
            
            status_str = str(result.get("Состояние", "")).strip()
            status_str = ''.join(c for c in status_str if ord(c) >= 32 or c in '\n\t\r')
            
            index_str = str(result.get("Индекс состояния", "")).strip()
            index_str = ''.join(c for c in index_str if ord(c) >= 32 or c in '\n\t\r')
            
            complaints_str = str(result.get("Количество жалоб", "0")).strip()
            complaints_str = ''.join(c for c in complaints_str if ord(c) >= 32 or c in '\n\t\r')
            
            unit_str = str(result.get("Подразделение", "")).strip()
            unit_str = ''.join(c for c in unit_str if ord(c) >= 32 or c in '\n\t\r')
            
            # Format as simple table row - PLAIN TEXT ONLY
            row = f"{idx} | {time_str:10} | {status_str:30} | {index_str:6} | {complaints_str:6} | {unit_str}"
            content.append(row)
        
        content.append("-" * 70)
        content.append("")
        content.append("Отчет сгенерирован автоматически")
        content.append(f"Дата создания: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
        
        # Write ONLY plain text with UTF-8 encoding
        # NO binary data, NO Excel structures, ONLY readable characters
        text_content = '\n'.join(content)
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(text_content)
        
        # Verify file was written correctly - check for binary markers
        with open(file_path, 'rb') as f:
            raw_bytes = f.read(100)
            # Check for Excel binary markers (PK zip header or OLE markers)
            if b'PK' in raw_bytes or b'\xd0\xcf' in raw_bytes:
                raise ValueError("Binary data detected - export failed!")
        
        return file_path, None
        
    except (OSError, ValueError, Exception) as exc:
        return None, f"Ошибка при создании отчёта: {str(exc)}"


def evaluate_condition(sensations_scores, complaints_list):
    """Evaluate condition based on collected scores and complaints
    Works with any set of sensations (randomized or fixed)"""
    
    raw_scores = {
        question: int(radio.value or 0)
        for question, radio in sensations_scores.items()
    }
    
    # Find matching categories in the current questions
    # This makes it flexible to work with any set of questions
    positive_keywords = ["хорошее", "уверенность", "спокойствие", "хорошее общее"]
    negative_keywords = ["плохое", "неуверенность", "безразличие", "напряжение", "страх"]
    exhaustion_keywords = ["усталость", "невнимательность"]
    
    positive_score = 0
    negative_score = 0
    exhaustion_score = 0
    
    for question, score in raw_scores.items():
        q_lower = question.lower()
        if any(kw in q_lower for kw in positive_keywords):
            positive_score += score
        elif any(kw in q_lower for kw in negative_keywords):
            negative_score += score
        elif any(kw in q_lower for kw in exhaustion_keywords):
            exhaustion_score += score
    
    # Adjust score: positive adds, negative subtracts (inverted), exhaustion penalizes
    adjusted_score = positive_score + (10 - negative_score) + max(0, 6 - exhaustion_score * 2)
    complaints_count = len(complaints_list)
    final_score = max(adjusted_score - min(complaints_count, 6), 0)

    # Build risk and resource markers
    risk_markers = []
    resource_markers = []
    
    for question, score in raw_scores.items():
        q_lower = question.lower()
        if score >= 1:
            if any(kw in q_lower for kw in ["усталость", "невнимательность", "затруднение"]):
                if "утомления" not in str(risk_markers):
                    risk_markers.append("есть признаки утомления")
            elif any(kw in q_lower for kw in ["плохое", "дискомфорт"]):
                if "дискомфорт" not in str(risk_markers):
                    risk_markers.append("заметен физический дискомфорт")
            elif any(kw in q_lower for kw in ["плохое настроение", "безразличие"]):
                if "эмоциональный" not in str(risk_markers):
                    risk_markers.append("эмоциональный фон снижен")
            elif any(kw in q_lower for kw in ["неуверенность", "страх"]):
                if "неустойчивость" not in str(risk_markers):
                    risk_markers.append("ощущается неустойчивость")
            elif any(kw in q_lower for kw in ["хорошее", "уверенность", "спокойствие"]):
                if score == 2:
                    resource_markers.append(question.lower())
    
    if complaints_count >= 3:
        if "телесный дискомфорт" not in " ".join(risk_markers):
            risk_markers.append("заметен телесный дискомфорт")

    # Determine status based on score
    if final_score >= 15 and complaints_count == 0 and negative_score <= 3:
        status_text = "Зелёный — ресурсное состояние"
        status_color = ft.Colors.GREEN
        summary = "Самочувствие устойчивое, выраженных признаков перегрузки не выявлено."
        recommendation = "Можно продолжать работу в обычном режиме, сохраняя базовый отдых и питьевой режим."
    elif final_score >= 11 and complaints_count <= 3:
        status_text = "Жёлтый — в целом стабильное состояние"
        status_color = ft.Colors.AMBER
        summary = "Состояние в целом удовлетворительное, но уже есть отдельные признаки напряжения."
        recommendation = "Полезны короткая пауза, вода, смена активности и контроль нагрузки."
    elif final_score >= 7:
        status_text = "Оранжевый — заметное утомление"
        status_color = ft.Colors.ORANGE
        summary = "Отмечаются усталость или эмоциональное снижение; ресурсу нужно восстановление."
        recommendation = "Желательны отдых, снижение интенсивности задач и повторная оценка состояния позже."
    else:
        status_text = "Красный — неблагоприятное состояние"
        status_color = ft.Colors.RED
        summary = "Признаки утомления и дискомфорта выражены; требуется больше внимания к самочувствию."
        recommendation = "Нужны отдых, снижение текущей нагрузки и, при необходимости, консультация специалиста."

    details = []
    if risk_markers:
        details.append("Риски: " + "; ".join(risk_markers) + ".")
    if resource_markers:
        details.append("Ресурсы: " + "; ".join(set(resource_markers[:3])) + ".")
    if not details:
        details.append("Выраженных маркеров напряжения не отмечено.")

    return {
        "status_text": status_text,
        "status_color": status_color,
        "summary": summary,
        "recommendation": recommendation,
        "details": details,
        "final_score": final_score,
        "positive_score": positive_score,
        "negative_score": negative_score,
        "complaints_count": complaints_count,
    }


def report_startup_error(exc: Exception):
    """Log startup error in a cross-platform way (no Windows-specific code)."""
    log_path = None

    log_path = None
    try:
        LOGS_DIR.mkdir(parents=True, exist_ok=True)
        log_path = LOGS_DIR / "startup_error.log"
        with log_path.open("a", encoding="utf-8") as log_file:
            log_file.write(f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]\n")
            log_file.write("".join(traceback.format_exception(type(exc), exc, exc.__traceback__)))
    except (OSError, TypeError, AttributeError):
        log_path = None

    message = (
        f"Ошибка при запуске приложения: {exc}\n"
        f"Время: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    if log_path:
        message += f"\nЛог сохранений в: {log_path}"

    # Cross-platform error reporting (no Windows-specific calls)
    print(f"STARTUP ERROR: {message}")
    try:
        import traceback as tb
        traceback.print_exc()
    except Exception:
        pass


def main(page: ft.Page):
    """Main application entry point with robust error handling for all platforms."""
    platform = "android"  # Default to android for safety
    
    try:
        print("DEBUG: main() called")
        try:
            platform = get_platform_name()
            print(f"DEBUG: Running on {platform}")
        except Exception as e:
            print(f"WARNING: Platform detection failed: {e}, defaulting to 'android'")
        
        # Ensure storage is initialized before anything else
        try:
            ensure_storage_dirs()
            print("DEBUG: storage dirs initialized")
        except Exception as e:
            print(f"WARNING: Failed to initialize storage: {e}")
        
        # Configure page with safe defaults
        try:
            page.title = APP_TITLE
            page.theme_mode = ft.ThemeMode.DARK
            page.theme = ft.Theme(color_scheme_seed=ft.Colors.BLUE_GREY_900)
            page.scroll = ft.ScrollMode.AUTO
            
            # Adaptive window sizing: desktop only
            if platform == "windows":
                try:
                    page.window_width = 470
                    page.window_height = 840
                    page.window_min_width = 360
                    page.window_min_height = 640
                except Exception as e:
                    print(f"WARNING: Could not set window size: {e}")
            
            # Safe padding and layout with white design
            try:
                page.padding = 12
                page.bgcolor = "#ffffff"  # White background
                page.horizontal_alignment = ft.CrossAxisAlignment.STRETCH
            except Exception as e:
                print(f"WARNING: Could not set page styling: {e}")
                
        except Exception as e:
            print(f"ERROR: Failed to configure page: {e}")
            traceback.print_exc()
            # Continue anyway - page may still be usable
    except Exception as e:
        print(f"ERROR: Critical error in page initialization: {e}")
        traceback.print_exc()

    user_data = {}
    sensations_scores = {}
    complaints_list = []
    latest_result = {}
    latest_saved_table = None
    latest_conclusion = ""
    
    # Current test session questions (randomized)
    current_sensations = []
    current_complaints = []

    def show_snack(message: str, color: str = ft.Colors.RED_700):
        try:
            page.snack_bar = ft.SnackBar(
                ft.Text(message, color=ft.Colors.WHITE),
                bgcolor=color,
            )
            page.snack_bar.open = True
            page.update()
        except Exception as e:
            # Fallback if SnackBar fails
            print(f"Notification: {message}")

    def describe_file_error(prefix: str, exc: Exception) -> str:
        message = str(exc)
        lowered = message.lower()
        if "permission" in lowered or "access is denied" in lowered:
            return f"{prefix}. Проверьте, не открыт ли файл в другой программе и есть ли доступ к папке."
        return f"{prefix}: {message}"

    def build_step_indicator(current_step: int):
        steps = [
            (1, "Данные"),
            (2, "Оценка"),
            (3, "Жалобы"),
            (4, "Итог"),
        ]
        controls = []
        for step_number, label in steps:
            is_active = step_number == current_step
            is_done = step_number < current_step
            
            controls.append(
                ft.Row([
                    ft.Container(
                        content=ft.Text(
                            str(step_number), 
                            color=ft.Colors.WHITE if (is_active or is_done) else "#9ca3af",
                            size=12, 
                            weight=ft.FontWeight.BOLD
                        ),
                        width=28,
                        height=28,
                        border_radius=14,
                        bgcolor="#000000" if (is_active or is_done) else "#e5e7eb",
                        alignment=ft.Alignment(0, 0),
                    ),
                    ft.Text(label, size=13, color="#6b7280", weight=ft.FontWeight.W_500),
                ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER)
            )
        
        return ft.Row(controls, spacing=12, wrap=True)

    def build_daily_trend_chart(rows):
        if not rows:
            return [ft.Text("Пока нет данных для построения графика.", color=ft.Colors.BLUE_GREY_700)]

        daily_scores = {}
        daily_counts = {}
        for row in rows:
            date_value = row.get("Дата", "")
            if not date_value:
                continue
            daily_scores[date_value] = daily_scores.get(date_value, 0) + parse_score_value(row.get("Индекс состояния", 0))
            daily_counts[date_value] = daily_counts.get(date_value, 0) + 1

        chart_controls = []
        for date_value in sorted(daily_scores.keys(), key=parse_date_value)[-10:]:
            avg_score = round(daily_scores[date_value] / max(daily_counts[date_value], 1), 1)
            if avg_score >= 15:
                color = ft.Colors.GREEN_600
            elif avg_score >= 11:
                color = ft.Colors.AMBER_700
            elif avg_score >= 7:
                color = ft.Colors.ORANGE_700
            else:
                color = ft.Colors.RED_600
            chart_controls.append(
                ft.Container(
                    content=ft.Column(
                        [
                            ft.Row(
                                [
                                    ft.Container(ft.Text(date_value, size=12), width=96),
                                    ft.Text(f"Средний индекс: {avg_score}/20", weight=ft.FontWeight.W_600),
                                ],
                                spacing=8,
                            ),
                            ft.Container(
                                content=ft.Container(
                                    height=10,
                                    width=max(18, min(240, int((avg_score / 20) * 240))),
                                    bgcolor=color,
                                    border_radius=999,
                                ),
                                bgcolor=ft.Colors.BLUE_GREY_50,
                                border_radius=999,
                                padding=2,
                            ),
                        ],
                        spacing=6,
                    ),
                    bgcolor=ft.Colors.WHITE,
                    border_radius=12,
                    padding=10,
                )
            )
        return chart_controls

    def open_history_secure(e=None):
        show_history()

    def open_backups_folder(e=None):
        try:
            ensure_storage_dirs()
            os.startfile(str(BACKUPS_DIR))
        except OSError as exc:
            show_snack(describe_file_error("Не удалось открыть папку резервных копий", exc))

    def create_manual_backup(e=None):
        try:
            file_path, error = create_results_backup("manual")
            if file_path:
                show_snack(f"Копия базы создана: {file_path}", ft.Colors.GREEN_700)
            else:
                show_snack(f"Не удалось создать копию базы: {error}")
        except OSError as exc:
            show_snack(describe_file_error("Не удалось создать резервную копию", exc))

    def export_all_history(e=None):
        history_rows = load_history_rows()
        file_path, error = export_history_to_excel(history_rows)
        if file_path:
            show_snack(f"История экспортирована: {file_path}", ft.Colors.GREEN_700)
        else:
            show_snack(f"Не удалось экспортировать историю: {error}")

    def print_current_report(e=None):
        if not latest_result:
            show_snack("Сначала сформируйте результат для печати.", ft.Colors.AMBER_700)
            return
        file_path, error = export_result_to_pdf(user_data, latest_result, complaints_list)
        if not file_path:
            show_snack(f"Не удалось подготовить PDF для печати: {error}")
            return
        try:
            os.startfile(str(file_path), "print")
            show_snack("Команда печати отправлена. Проверьте принтер и очередь печати.", ft.Colors.GREEN_700)
        except OSError as exc:
            show_snack(describe_file_error("Не удалось отправить отчёт на печать", exc))

    def copy_current_conclusion(e=None):
        nonlocal latest_conclusion
        if not latest_conclusion:
            show_snack("Сначала сформируйте результат, чтобы скопировать заключение.", ft.Colors.AMBER_700)
            return
        try:
            page.set_clipboard(latest_conclusion)
            show_snack("Краткое заключение скопировано в буфер обмена.", ft.Colors.GREEN_700)
        except Exception as exc:
            show_snack(describe_file_error("Не удалось скопировать заключение", exc))

    def build_section_header(title: str, subtitle: str, icon: str = ft.Icons.MEDICAL_SERVICES_ROUNDED):
        return ft.Container(
            content=ft.Column(
                [
                    ft.Text(title, size=24, weight=ft.FontWeight.BOLD, color="#1f2937"),
                    ft.Text(subtitle, size=12, color="#6b7280"),
                ],
                spacing=2,
            ),
            padding=16,
            margin=ft.margin.only(bottom=16),
        )

    def build_metric_card(title: str, value: str, icon: str, color: str):
        return ft.Column(
            [
                ft.Text(title, size=11, color="#9ca3af", weight=ft.FontWeight.W_500),
                ft.Text(str(value), size=18, weight=ft.FontWeight.BOLD, color="#1f2937"),
            ],
            spacing=2,
        )

    def build_info_tile(label: str, value: str):
        return ft.Column(
            [
                ft.Text(label, size=11, color="#9ca3af"),
                ft.Text(value or "—", size=14, weight=ft.FontWeight.W_500, color="#1f2937"),
            ],
            spacing=2,
        )

    def open_results_folder(e=None):
        try:
            ensure_storage_dirs()
            platform = get_platform_name()
            if platform == "windows":
                os.startfile(str(RESULTS_DIR))
            elif platform == "osx":
                os.system(f'open "{RESULTS_DIR}"')
            elif platform == "linux":
                os.system(f'xdg-open "{RESULTS_DIR}"')
            else:
                # Android - просто уведомим пользователя
                show_snack(f"Папка результатов: {RESULTS_DIR}", ft.Colors.BLUE_700)
        except Exception as exc:
            show_snack(f"Не удалось открыть папку: {exc}")

    def open_exports_folder(e=None):
        try:
            ensure_storage_dirs()
            platform = get_platform_name()
            if platform == "windows":
                os.startfile(str(EXPORTS_DIR))
            elif platform == "osx":
                os.system(f'open "{EXPORTS_DIR}"')
            elif platform == "linux":
                os.system(f'xdg-open "{EXPORTS_DIR}"')
            else:
                # Android - просто уведомим пользователя
                show_snack(f"Папка экспорта: {EXPORTS_DIR}", ft.Colors.BLUE_700)
        except Exception as exc:
            show_snack(f"Не удалось открыть папку экспорта: {exc}")

    def create_daily_report(e=None):
        """Generate and export daily duty report (4+ test sessions)."""
        if not user_data or not user_data.get("fio"):
            show_snack("Требуются личные данные для создания отчёта.", ft.Colors.AMBER_700)
            return
        
        # Generate report
        report_data, error = generate_daily_report(user_data)
        
        if error:
            show_snack(f"⚠️ {error}", ft.Colors.ORANGE_700)
            return
        
        # Export to Excel
        file_path, export_error = export_daily_report(report_data)
        
        if file_path:
            show_snack(
                f"✅ Суточный отчет создан!\n{Path(file_path).name}",
                ft.Colors.GREEN_700
            )
            
            # Try to open the file
            try:
                if sys.platform == "win32":
                    # Windows: use os.startfile()
                    os.startfile(str(file_path))
                elif sys.platform == "darwin":
                    # macOS: use open command
                    subprocess.Popen(["open", str(file_path)])
                else:
                    # Linux: try xdg-open
                    subprocess.Popen(["xdg-open", str(file_path)])
            except Exception as open_error:
                print(f"DEBUG: Could not auto-open file: {open_error}")
                # File was still created successfully, just couldn't open it
                show_snack(f"Файл создан, но не удалось открыть. Откройте вручную: {Path(file_path).name}", ft.Colors.BLUE_700)
        else:
            show_snack(f"❌ Ошибка сохранения: {export_error}", ft.Colors.RED_600)

    def export_current_excel(e=None):
        if not latest_result:
            show_snack("Сначала сформируйте результат.", ft.Colors.AMBER_700)
            return
        file_path, error = export_result_to_excel(user_data, latest_result, complaints_list)
        if file_path:
            show_snack(f"Excel сохранён: {file_path}", ft.Colors.GREEN_700)
        else:
            show_snack(f"Не удалось сохранить Excel: {error}")

    def export_current_pdf(e=None):
        if not latest_result:
            show_snack("Сначала сформируйте результат.", ft.Colors.AMBER_700)
            return
        file_path, error = export_result_to_pdf(user_data, latest_result, complaints_list)
        if file_path:
            show_snack(f"PDF сохранён: {file_path}", ft.Colors.GREEN_700)
        else:
            show_snack(f"Не удалось сохранить PDF: {error}")

    def show_history(e=None):
        page.clean()
        try:
            history_rows = load_history_rows()
        except Exception as ex:
            print(f"ERROR: Could not load history: {ex}")
            page.add(ft.Text(f"⚠️ Ошибка загрузки: {str(ex)[:50]}"))
            return
            
        fio_filter = ft.TextField(
            label="Ф.И.О.",
            prefix_icon=ft.Icons.PERSON_SEARCH_ROUNDED,
        )
        unit_filter = ft.TextField(
            label="Подразделение",
            prefix_icon=ft.Icons.GROUP_WORK_ROUNDED,
        )
        date_filter = ft.TextField(
            label="Дата",
            hint_text="ДД.ММ.ГГГГ",
            prefix_icon=ft.Icons.CALENDAR_MONTH_ROUNDED,
        )
        status_filter = ft.Dropdown(
            label="Уровень состояния",
            options=[
                ft.dropdown.Option(text="Все", value="Все"),
                ft.dropdown.Option(text="Зелёный", value="Зелёный"),
                ft.dropdown.Option(text="Жёлтый", value="Жёлтый"),
                ft.dropdown.Option(text="Оранжевый", value="Оранжевый"),
                ft.dropdown.Option(text="Красный", value="Красный"),
            ],
            value="Все",
        )
        summary_text = ft.Text(color=ft.Colors.BLUE_GREY_700)
        stats_wrap = ft.Column(spacing=8)
        trend_column = ft.Column(spacing=8)
        history_list = ft.Column(spacing=8)

        def refresh_history(e=None):
            fio_query = (fio_filter.value or "").strip().lower()
            unit_query = (unit_filter.value or "").strip().lower()
            date_query = (date_filter.value or "").strip().lower()
            selected_status = status_filter.value or "Все"
            filtered = []

            for row in history_rows:
                if fio_query and fio_query not in row.get("Ф.И.О.", "").lower():
                    continue
                if unit_query and unit_query not in row.get("Подразделение", "").lower():
                    continue
                if date_query and date_query not in row.get("Дата", "").lower():
                    continue
                if selected_status != "Все" and not normalize_status_name(row.get("Состояние", "")).startswith(selected_status):
                    continue
                filtered.append(row)

            counts = Counter(normalize_status_name(row.get("Состояние", "")) for row in filtered)
            summary_text.value = (
                f"Найдено записей: {len(filtered)}"
                + (f" • Последняя запись: {filtered[0].get('Дата', '')} {filtered[0].get('Время', '')}" if filtered else "")
            )
            stats_wrap.controls = [
                build_metric_card("Зелёный", counts.get("Зелёный", 0), ft.Icons.CHECK_CIRCLE_ROUNDED, ft.Colors.GREEN_600),
                build_metric_card("Жёлтый", counts.get("Жёлтый", 0), ft.Icons.INFO_ROUNDED, ft.Colors.AMBER_700),
                build_metric_card("Оранжевый", counts.get("Оранжевый", 0), ft.Icons.WARNING_AMBER_ROUNDED, ft.Colors.ORANGE_700),
                build_metric_card("Красный", counts.get("Красный", 0), ft.Icons.ERROR_ROUNDED, ft.Colors.RED_600),
            ]
            trend_column.controls = build_daily_trend_chart(filtered)

            history_list.controls.clear()
            if not filtered:
                history_list.controls.append(
                    ft.Container(
                        content=ft.Text("Записи не найдены. Уточните фильтры или сначала заполните анкету."),
                        bgcolor=ft.Colors.WHITE,
                        border_radius=14,
                        padding=12,
                    )
                )
            else:
                for row in filtered[:150]:
                    color = get_status_color(row.get("Состояние", ""))
                    history_list.controls.append(
                        ft.Container(
                            content=ft.Column(
                                [
                                    ft.Row(
                                        [
                                            ft.Text(row.get("Ф.И.О.", "Без имени"), weight=ft.FontWeight.BOLD, expand=True),
                                            ft.Text(
                                                f"{row.get('Дата', '')} {row.get('Время', '')}".strip(),
                                                color=ft.Colors.BLUE_GREY_700,
                                            ),
                                        ]
                                    ),
                                    ft.Text(row.get("Состояние", ""), color=color, weight=ft.FontWeight.W_600),
                                    ft.Text(f"Подразделение: {row.get('Подразделение', '—')}", size=13),
                                    ft.Text(f"Дежурство: {row.get('Вид дежурства', '—')}", size=13),
                                    ft.Text(
                                        f"Индекс: {row.get('Индекс состояния', '—')} | Жалобы: {row.get('Количество жалоб', '—')}",
                                        size=12,
                                        color=ft.Colors.BLUE_GREY_700,
                                    ),
                                ],
                                spacing=4,
                            ),
                            bgcolor=ft.Colors.WHITE,
                            border=ft.border.all(1, ft.Colors.with_opacity(0.25, color)),
                            border_radius=14,
                            padding=10,
                        )
                    )
                if len(filtered) > 150:
                    history_list.controls.append(
                        ft.Text(
                            f"Показаны первые 150 записей из {len(filtered)}.",
                            italic=True,
                            color=ft.Colors.BLUE_GREY_700,
                        )
                    )
            page.update()

        fio_filter.on_change = refresh_history
        unit_filter.on_change = refresh_history
        date_filter.on_change = refresh_history
        status_filter.on_change = refresh_history

        page.add(
            build_section_header("История результатов", "Защищённый раздел психолога: график, фильтры и экспорт", ft.Icons.HISTORY_ROUNDED),
            build_step_indicator(4),

            ft.Container(
                content=ft.Column(
                    [
                        ft.Text("Фильтры", size=16, weight=ft.FontWeight.BOLD),
                        ft.Column(
                            spacing=8,
                            controls=[fio_filter, unit_filter, date_filter, status_filter],
                        ),
                        summary_text,
                    ],
                    spacing=8,
                ),
                bgcolor=ft.Colors.WHITE,
                border_radius=14,
                padding=12,
            ),
            ft.Text("График состояния по датам", size=17, weight=ft.FontWeight.BOLD),
            ft.Container(content=trend_column, bgcolor=ft.Colors.WHITE, border_radius=14, padding=12),
            stats_wrap,
            ft.Text("Журнал наблюдений", size=17, weight=ft.FontWeight.BOLD),
            history_list,
            ft.Column(
                spacing=8,
                controls=[
                    ft.ElevatedButton("Экспорт всей истории в Excel", icon=ft.Icons.TABLE_VIEW_ROUNDED, on_click=export_all_history),
                    ft.OutlinedButton("Создать копию базы результатов", icon=ft.Icons.BACKUP_ROUNDED, on_click=create_manual_backup),
                    ft.OutlinedButton("Открыть резервные копии", icon=ft.Icons.FOLDER_COPY_ROUNDED, on_click=open_backups_folder),
                    ft.OutlinedButton("Назад на главную", icon=ft.Icons.HOME_ROUNDED, on_click=lambda _: show_step_1()),
                ],
            ),
        )
        refresh_history()

    def show_step_1(e=None):
        try:
            page.clean()

            fio = ft.TextField(label="Ф.И.О.", prefix_icon=ft.Icons.PERSON_ROUNDED, filled=True)
            rank = ft.TextField(label="В/звание", prefix_icon=ft.Icons.BADGE_ROUNDED, filled=True)
            unit = ft.TextField(label="Подразделение", prefix_icon=ft.Icons.GROUP_WORK_ROUNDED, filled=True)
            duty = ft.TextField(label="Вид дежурства", prefix_icon=ft.Icons.ASSIGNMENT_ROUNDED, filled=True)
            date = ft.TextField(
                label="Дата",
                value=get_current_date_text(),
                prefix_icon=ft.Icons.CALENDAR_MONTH_ROUNDED,
                read_only=True,
                filled=True,
            )

            page.add(
                build_section_header(APP_TITLE, APP_SUBTITLE),
                build_step_indicator(1),
                ft.Divider(height=1, color="#e5e7eb"),
                ft.Text("I. Личные данные", size=16, weight=ft.FontWeight.BOLD, color="#1f2937"),
                ft.Text(
                    "Заполните обязательные поля.",
                    size=12,
                    color="#6b7280",
                ),
                fio,
                rank,
                unit,
                duty,
                date,
                ft.Container(height=16),
                ft.ElevatedButton(
                    "Продолжить к оценке",
                    icon=ft.Icons.ARROW_FORWARD_ROUNDED,
                    on_click=lambda _: save_step_1(fio, rank, unit, duty, date),
                    width=300,
                ),
                ft.Container(height=8),
                ft.OutlinedButton("История результатов", icon=ft.Icons.HISTORY_ROUNDED, on_click=open_history_secure, width=300),
                ft.OutlinedButton("Создать резервную копию", icon=ft.Icons.BACKUP_ROUNDED, on_click=create_manual_backup, width=300),
                ft.OutlinedButton("Открыть папку результатов", icon=ft.Icons.FOLDER_OPEN_ROUNDED, on_click=open_results_folder, width=300),
            )
            page.update()
        except Exception as e:
            print(f"ERROR in show_step_1: {e}")
            traceback.print_exc()

    def save_step_1(fio_field, rank_field, unit_field, duty_field, date_field):
        required_fields = [
            (fio_field, "Ф.И.О."),
            (rank_field, "В/звание"),
            (unit_field, "Подразделение"),
            (duty_field, "Вид дежурства"),
        ]
        missing = []

        for control, label in required_fields:
            control.error_text = None
            if not (control.value or "").strip():
                control.error_text = "Обязательное поле"
                missing.append(label)

        date_field.error_text = None
        date_value = (date_field.value or "").strip()
        if not date_value or parse_date_value(date_value) == datetime.min:
            date_value = get_current_date_text()
            date_field.value = date_value

        if missing:
            page.update()
            show_snack(f"Проверьте обязательные поля: {', '.join(missing)}")
            return

        user_data.update(
            {
                "fio": fio_field.value.strip(),
                "rank": rank_field.value.strip(),
                "unit": unit_field.value.strip(),
                "duty": duty_field.value.strip(),
                "date": date_value,
            }
        )
        show_step_2()

    def show_step_2():
        nonlocal current_sensations
        page.clean()
        sensations_scores.clear()
        
        # Generate random sensations for this session
        if QUESTIONS_BANK:
            current_sensations = get_random_questions("sensations")
        else:
            current_sensations = SENSATIONS.copy()
        
        # Shuffle the order
        random.shuffle(current_sensations)
        
        page.add(
            build_section_header("Оценка ощущений", "Отметьте выраженность каждого состояния", ft.Icons.MONITOR_HEART_ROUNDED),
            build_step_indicator(2),
            ft.Text("II. Оцените ощущения", size=18, weight=ft.FontWeight.BOLD, color="#1f2937"),
            ft.Text(f"2 — сильно, 1 — слабо, 0 — нейтрально (вопросов: {len(current_sensations)})", italic=True, size=13, color="#6b7280"),
        )

        for question in current_sensations:
            radio_group = ft.RadioGroup(
                content=ft.Row(
                    [
                        ft.Radio(value="0", label="☐ 0", active_color="#000000", label_style=ft.TextStyle(color="#1f2937", size=13)),
                        ft.Radio(value="1", label="☐ 1", active_color="#000000", label_style=ft.TextStyle(color="#1f2937", size=13)),
                        ft.Radio(value="2", label="☐ 2", active_color="#000000", label_style=ft.TextStyle(color="#1f2937", size=13)),
                    ],
                    spacing=20,
                ),
                value="0",
            )
            sensations_scores[question] = radio_group
            page.add(
                ft.Column([
                    ft.Text(question, weight=ft.FontWeight.W_600, color="#1f2937", size=14),
                    ft.Container(
                        content=radio_group,
                        padding=12,
                        bgcolor="#f9fafb",
                        border_radius=8,
                    ),
                ], spacing=6)
            )

        page.add(
            ft.Column(
                spacing=8,
                controls=[
                    ft.ElevatedButton("Далее к жалобам", icon=ft.Icons.ARROW_FORWARD_ROUNDED, on_click=lambda _: show_step_3()),
                    ft.OutlinedButton("Назад", icon=ft.Icons.ARROW_BACK_ROUNDED, on_click=lambda _: show_step_1()),
                ],
            )
        )
        page.update()

    def show_step_3():
        nonlocal current_complaints
        page.clean()
        
        # Generate random complaints for this session
        if QUESTIONS_BANK:
            current_complaints = get_random_questions("symptoms")
        else:
            current_complaints = COMPLAINTS.copy()
        
        # Shuffle the order
        random.shuffle(current_complaints)
        
        page.add(
            build_section_header("Жалобы", "При необходимости отметьте наблюдаемые симптомы", ft.Icons.HEALTH_AND_SAFETY_ROUNDED),
            build_step_indicator(3),
            ft.Divider(height=1, color="#e5e7eb"),
            ft.Text("III. Отметьте жалобы", size=16, weight=ft.FontWeight.BOLD, color="#1f2937"),
            ft.Text(f"Можно оставить пустым, если жалоб нет. (вопросов: {len(current_complaints)})", size=12, color="#6b7280"),
        )

        checkboxes = []
        for complaint in current_complaints:
            checkbox = ft.Checkbox(label=complaint, value=complaint in complaints_list, label_style=ft.TextStyle(color="#1f2937", size=13))
            checkboxes.append(checkbox)
            page.add(checkbox)

        def save_and_finish(e):
            complaints_list.clear()
            for checkbox in checkboxes:
                if checkbox.value:
                    complaints_list.append(checkbox.label)
            show_result()

        page.add(
            ft.Column(
                spacing=8,
                controls=[
                    ft.ElevatedButton("Показать результат", icon=ft.Icons.INSIGHTS_ROUNDED, on_click=save_and_finish),
                    ft.OutlinedButton("Назад", icon=ft.Icons.ARROW_BACK_ROUNDED, on_click=lambda _: show_step_2()),
                ],
            )
        )
        page.update()

    def show_result():
        nonlocal latest_saved_table
        try:
            page.clean()
            user_data["time"] = get_current_time_text()
            
            try:
                result = evaluate_condition(sensations_scores, complaints_list)
            except Exception as e:
                print(f"ERROR: Failed to evaluate condition: {e}")
                traceback.print_exc()
                show_snack(f"Ошибка расчета результата: {e}")
                show_step_1()
                return
            
            try:
                saved_table, save_error = save_result_to_table(user_data, result, complaints_list)
            except Exception as e:
                print(f"ERROR: Failed to save result: {e}")
                traceback.print_exc()
                saved_table, save_error = None, str(e)
            
            latest_result.clear()
            latest_result.update(result)
            latest_saved_table = saved_table
            
            # Display results section
            page.add(
                build_section_header("Результат", "Ваша интегральная оценка состояния", ft.Icons.ASSESSMENT_ROUNDED),
                build_step_indicator(4),
            )
            
            # Extract status info
            status_text = result.get("status_text", "Не определено")
            status_color = result.get("status_color", ft.Colors.GREY)
            summary = result.get("summary", "")
            recommendation = result.get("recommendation", "")
            details = result.get("details", [])
            final_score = result.get("final_score", 0)
            complaints_count = result.get("complaints_count", 0)
            
            # Build color indicator UI
            color_name = "Зелёный" if "Зелёный" in status_text else \
                        "Жёлтый" if "Жёлтый" in status_text else \
                        "Оранжевый" if "Оранжевый" in status_text else "Красный"
            
            page.add(
                ft.Divider(height=1, color="#e5e7eb"),
                ft.Column([
                    # Color indicator
                    ft.Row([
                        ft.Container(
                            width=80,
                            height=80,
                            bgcolor=status_color,
                            border_radius=16,
                        ),
                        ft.Column([
                            ft.Text(color_name, size=24, weight=ft.FontWeight.BOLD, color=status_color),
                            ft.Text(status_text.replace(f"{color_name} — ", ""), size=13, color="#6b7280"),
                        ], spacing=4),
                    ], spacing=16),
                    ft.Container(height=8),
                    ft.Text(summary, size=13, color="#374151"),
                    ft.Container(height=16),
                    
                    # Score metrics
                    ft.Row([
                        ft.Column([
                            ft.Text("Балл", size=11, weight=ft.FontWeight.W_600, color="#6b7280"),
                            ft.Text(str(final_score), size=18, weight=ft.FontWeight.BOLD, color="#1f2937"),
                        ], spacing=2, alignment=ft.MainAxisAlignment.CENTER),
                        ft.VerticalDivider(width=1, color="#e5e7eb"),
                        ft.Column([
                            ft.Text("Жалобы", size=11, weight=ft.FontWeight.W_600, color="#6b7280"),
                            ft.Text(str(complaints_count), size=18, weight=ft.FontWeight.BOLD, color="#1f2937"),
                        ], spacing=2, alignment=ft.MainAxisAlignment.CENTER),
                    ], expand=True, spacing=16),
                    ft.Container(height=16),
                    
                    # Brief analysis
                    ft.Text("Краткий анализ", size=13, weight=ft.FontWeight.W_600, color="#1f2937"),
                    ft.Column([
                        ft.Text(f"• {item}", size=12, color="#374151")
                        for item in details
                    ], spacing=6),
                    ft.Container(height=16),
                    
                    # Recommendation
                    ft.Text("Рекомендация", size=13, weight=ft.FontWeight.W_600, color="#1f2937"),
                    ft.Text(recommendation, size=12, color="#374151", selectable=True),
                    ft.Container(height=12),
                    
                    # Save confirmation
                    ft.Text("✓ Результат сохранён", size=12, color="#059669", weight=ft.FontWeight.W_500),
                ], spacing=0)
            )
            
            page.add(
                ft.Column(
                    spacing=8,
                    controls=[
                        ft.ElevatedButton(
                            "Пройти тест заново",
                            icon=ft.Icons.REFRESH_ROUNDED,
                            on_click=lambda e: show_step_1(),
                            width=300,
                        ),
                        ft.OutlinedButton(
                            "Создать суточный отчет (4 проведения)",
                            icon=ft.Icons.ASSESSMENT_ROUNDED,
                            on_click=create_daily_report,
                            width=300,
                        ),
                        ft.OutlinedButton(
                            "Открыть папку результатов",
                            icon=ft.Icons.FOLDER_OPEN_ROUNDED,
                            on_click=open_results_folder,
                            width=300,
                        ),
                    ],
                )
            )
        except Exception as e:
            print(f"ERROR in show_result: {e}")
            traceback.print_exc()
            show_snack(f"Ошибка отображения результата: {e}")
            show_step_1()

            save_message = (
                f"Результат добавлен в таблицу: {saved_table}"
                if saved_table
                else f"Не удалось сохранить таблицу: {save_error}"
            )
            save_color = ft.Colors.GREEN_700 if saved_table else ft.Colors.RED_700

            page.add(
                build_section_header("Результаты анализа", "Итоговая оценка текущего состояния", get_status_icon(result["status_text"])),
            ft.Container(
                content=ft.Column(
                    [
                        ft.Text(
                            f"Состояние: {result['status_text']}",
                            size=20,
                            weight=ft.FontWeight.BOLD,
                            color=ft.Colors.WHITE,
                        ),
                        ft.Text(result["summary"], size=14, color=ft.Colors.WHITE),
                    ],
                    spacing=6,
                ),
                bgcolor=result["status_color"],
                padding=14,
                border_radius=16,
            ),
            ft.Row(
                [
                    ft.Container(build_metric_card("Индекс состояния", f"{result['final_score']} / 20", ft.Icons.STACKED_LINE_CHART_ROUNDED, get_status_color(result["status_text"])), expand=True),
                    ft.Container(build_metric_card("Жалобы", result["complaints_count"], ft.Icons.HEALING_ROUNDED, ft.Colors.RED_500), expand=True),
                ],
                spacing=8,
            ),
            build_metric_card("Баланс маркеров", f"+{result['positive_score']} / -{result['negative_score']}", ft.Icons.BALANCE_ROUNDED, ft.Colors.BLUE_600),
            ft.Text(save_message, color=save_color),
            build_info_tile("Ф.И.О.", user_data.get("fio", "")),
            build_info_tile("Дата", user_data.get("date", "")),
            build_info_tile("Время прохождения", user_data.get("time", "")),
            build_info_tile("Звание", user_data.get("rank", "")),
            build_info_tile("Подразделение", user_data.get("unit", "")),
            build_info_tile("Вид дежурства", user_data.get("duty", "")),
            ft.Container(
                content=ft.Column(
                    [
                        ft.Text("Краткий анализ", size=18, weight=ft.FontWeight.BOLD),
                        ft.Column([ft.Text(f"• {item}") for item in result["details"]], spacing=6),
                        ft.Text("Рекомендация", size=18, weight=ft.FontWeight.BOLD),
                        ft.Text(result["recommendation"]),
                        ft.Text("Отмеченные жалобы", size=18, weight=ft.FontWeight.BOLD),
                        ft.Column(
                            [ft.Text(f"- {item}") for item in complaints_list] if complaints_list else [ft.Text("Жалобы не отмечены")],
                            spacing=4,
                        ),
                    ],
                    spacing=10,
                ),
                bgcolor=ft.Colors.WHITE,
                border_radius=16,
                padding=14,
            ),
            ft.Column(
                [
                    ft.ElevatedButton("Экспорт в Excel", icon=ft.Icons.TABLE_VIEW_ROUNDED, on_click=export_current_excel),
                    ft.ElevatedButton("Экспорт в PDF", icon=ft.Icons.PICTURE_AS_PDF_ROUNDED, on_click=export_current_pdf),
                    ft.OutlinedButton("История результатов", icon=ft.Icons.HISTORY_ROUNDED, on_click=show_history),
                    ft.OutlinedButton("Открыть папку экспорта", icon=ft.Icons.DOWNLOAD_ROUNDED, on_click=open_exports_folder),
                    ft.OutlinedButton("Открыть папку с результатами", icon=ft.Icons.FOLDER_OPEN_ROUNDED, on_click=open_results_folder),
                    ft.ElevatedButton("Пройти заново", icon=ft.Icons.RESTART_ALT_ROUNDED, on_click=lambda _: show_step_1()),
                ],
                spacing=8,
            ),
        )
            page.update()
        except Exception as e:
            print(f"ERROR in show_result: {e}")
            traceback.print_exc()
            try:
                page.clean()
                page.add(
                    ft.Column(
                        [
                            ft.Text("Ошибка при обработке результатов", size=20, weight=ft.FontWeight.BOLD, color=ft.Colors.RED),
                            ft.Text(f"Детали: {str(e)}", size=14),
                            ft.ElevatedButton("На главную", icon=ft.Icons.HOME_ROUNDED, on_click=lambda _: show_step_1()),
                        ],
                        spacing=12,
                    )
                )
                page.update()
            except Exception as e2:
                print(f"ERROR: Failed to display error message: {e2}")

    # Appbar disabled due to Flet compatibility issues
    # page.appbar = ft.AppBar(
    #     title=ft.Text(APP_TITLE, weight=ft.FontWeight.BOLD),
    #     bgcolor=ft.Colors.WHITE,
    #     actions=[
    #         ft.IconButton(ft.Icons.HOME_ROUNDED, tooltip="Главная", on_click=lambda _: show_step_1()),
    #         ft.IconButton(ft.Icons.HISTORY_ROUNDED, tooltip="История", on_click=show_history),
    #     ],
    # )

    # Initial screen setup
    try:
        print("DEBUG: Calling show_step_1() to initialize UI")
        show_step_1()
        print("DEBUG: show_step_1() initialization complete")
    except Exception as e:
        print(f"ERROR: Failed to initialize UI: {e}")
        traceback.print_exc()
        # Show fallback error message on page
        page.clean()
        page.add(
            ft.Column(
                [
                    ft.Text("Ошибка инициализации приложения", size=20, weight=ft.FontWeight.BOLD, color=ft.Colors.RED),
                    ft.Text(f"Детали: {str(e)}", size=14, color=ft.Colors.RED_600),
                    ft.Text("Пожалуйста, перезапустите приложение", size=12),
                ],
                spacing=12,
            )
        )
        page.update()


if __name__ == "__main__":
    import signal
    
    def signal_handler(sig, frame):
        """Handle shutdown signals gracefully on all platforms."""
        print('[AnketaSamootsenki] Shutting down...')
        try:
            sys.exit(0)
        except:
            pass
    
    # Register signal handlers safely (may fail on some platforms)
    try:
        if hasattr(signal, 'SIGINT'):
            signal.signal(signal.SIGINT, signal_handler)
        if hasattr(signal, 'SIGTERM'):
            signal.signal(signal.SIGTERM, signal_handler)
    except (ValueError, OSError) as e:
        print(f"WARNING: Could not register signal handlers: {e}")
    
    try:
        print("[AnketaSamootsenki] Starting application...")
        # Try Desktop mode first (requires Flet runtime)
        # If fails, fall back to WEB mode (built-in local server, no internet needed)
        try:
            ft.app(target=main)
        except Exception as desktop_error:
            if "URLError" in str(type(desktop_error)) or "getaddrinfo" in str(desktop_error):
                print(f"[AnketaSamootsenki] Desktop mode failed (no internet): {desktop_error}")
                print("[AnketaSamootsenki] Switching to WEB mode (built-in server)...")
                # WEB mode uses built-in HTTP server, no external downloads needed
                ft.app(target=main, port=8550)
            else:
                raise
    except Exception as exc:
        print(f"[AnketaSamootsenki] Application failed: {exc}")
        traceback.print_exc()
        report_startup_error(exc)
